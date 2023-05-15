
import numpy as np
import random
import pandas as pd
from scipy import stats
import openpyxl as xl


random.seed(5)

P = [33,101,151] #TODO establecer proveedores
E = [25,49,73,95] #TODO establecer etapas de decisión
indices = ['Fi compra','Fi venta','P gen','P carga','P compra max','P venta max','Bat cap']

I0 = 2  

for prosumers in P:
    for etapas in E:

        #CONJUNTOS
        N = [i for i in range(1,prosumers)]
        T = [i for i in range(1,etapas)]
        S = [i for i in range(3)]

        wb = xl.Workbook()
        del wb['Sheet']
        
        for c,i  in enumerate(indices,start=1):
            wb.create_sheet(index=c,title = indices[c-1])

        hoja = wb['Fi compra']
        hoja.cell(1,1).value = 'Proveedor'
        hoja.cell(1,2).value = 'Hora'
        hoja.cell(1,3).value = 'Escenario'
        fi_compra = {(i,t):[random.randint(160,200) for k in range(3)] for i in N for t in T}
        contador = 2
        for key in (fi_compra.keys()):
            for n,esc in enumerate(fi_compra[key],start=1):
                hoja.cell(contador,1).value = key[0]
                hoja.cell(contador,2).value = key[1]
                hoja.cell(contador,3).value = n
                hoja.cell(contador,4).value = fi_compra[key][n-1]
                contador += 1

        hoja = wb['Fi venta']
        hoja.cell(1,1).value = 'Proveedor'
        hoja.cell(1,2).value = 'Hora'
        hoja.cell(1,3).value = 'Escenario'
        fi_venta = {(i,t):[random.randint(100,110)*-1 for k in range(3)] for i in N for t in T}
        contador = 2
        for key in (fi_venta.keys()):
            for n,esc in enumerate(fi_venta[key],start=1):
                hoja.cell(contador,1).value = key[0]
                hoja.cell(contador,2).value = key[1]
                hoja.cell(contador,3).value = n
                hoja.cell(contador,4).value = fi_venta[key][n-1]
                contador += 1


        hoja = wb['P gen']
        hoja.cell(1,1).value = 'Proveedor'
        hoja.cell(1,2).value = 'Hora'
        hoja.cell(1,3).value = 'Escenario'
        p_gen = {(i,t):[random.randint(4,7) for k in range(3)] for i in N for t in T}       
        contador = 2
        for key in (p_gen.keys()):
            for n,esc in enumerate(p_gen[key],start=1):
                hoja.cell(contador,1).value = key[0]
                hoja.cell(contador,2).value = key[1]
                hoja.cell(contador,3).value = n
                hoja.cell(contador,4).value = p_gen[key][n-1]
                contador += 1

        hoja = wb['P carga']
        hoja.cell(1,1).value = 'Proveedor'
        hoja.cell(1,2).value = 'Hora'
        hoja.cell(1,3).value = 'Escenario'
        p_carga = {(i,t):[random.randint(6,10) for k in range(3)] for i in N for t in T}
        contador = 2
        for key in (p_carga.keys()):
            for n,esc in enumerate(p_carga[key],start=1):
                hoja.cell(contador,1).value = key[0]
                hoja.cell(contador,2).value = key[1]
                hoja.cell(contador,3).value = n
                hoja.cell(contador,4).value = p_carga[key][n-1]
                contador += 1

        hoja = wb['P compra max']
        hoja.cell(1,1).value = 'Proveedor'
        hoja.cell(1,2).value = 'Hora'
        max_p_compra = {(i,t):random.randint(10,20) for i in N for t in T}
        for n,key in enumerate(max_p_compra.keys(),start=2):
            hoja.cell(n,1).value = key[0]
            hoja.cell(n,2).value = key[1]
            hoja.cell(n,3).value = max_p_compra[key]

        hoja = wb['P venta max']
        hoja.cell(1,1).value = 'Proveedor'
        hoja.cell(1,2).value = 'Hora'
        max_p_venta = {(i,t):random.randint(10,20) for i in N for t in T}
        for n,key in enumerate(max_p_venta.keys(),start=2):
            hoja.cell(n,1).value = key[0]
            hoja.cell(n,2).value = key[1]
            hoja.cell(n,3).value = max_p_venta[key]

        hoja = wb['Bat cap']
        hoja.cell(1,1).value = 'Proveedor'
        k = {i:random.randint(70,90) for i in N}
        for n,key in enumerate(k.keys(),start=2):
            hoja.cell(n,1).value = key
            hoja.cell(n,2).value = k[key]

        hoja.cell(1,4).value = 'Inventario inicial prosumers'
        hoja.cell(1,5).value = I0


        wb.save(f'Instancia {prosumers-1} prosumers_{etapas-1} horas.xlsx')

